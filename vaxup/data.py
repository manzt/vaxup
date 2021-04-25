import re
from datetime import datetime
from enum import Enum
from itertools import groupby
from typing import Iterable, Literal, Optional, TYPE_CHECKING

from openpyxl import load_workbook
from pydantic import EmailStr, validator
from pydantic.types import PositiveInt

# Improve intellisense for VSCode
# https://github.com/microsoft/python-language-server/issues/1898
if TYPE_CHECKING:
    from dataclasses import dataclass
else:
    from pydantic.dataclasses import dataclass as dataclass

COLUMNS = {
    "Start Time": "start_time",
    "First Name": "first_name",
    "Last Name": "last_name",
    "Phone": "phone",
    "Email": "email",
    "Calendar": "location",
    "Date of birth (M/DD/YYYY)": "dob",
    "Street address (e.g., 60 Madison Ave.)": "street_address",
    "Apt / suite number": "apt",
    "City (e.g., Queens)": "city",
    "State (e.g., NY)": "state",
    "Zip code (e.g., 10010)": "zip_code",
    "Which race do you identify as?": "race",
    "Do you identify as Hispanic, Latino, or Latina?": "ethnicity",
    "What sex were you assigned at birth?": "sex",
    "COVID-19 vaccines are free for you and we will not be billing your insurance. For informational purposes, do you have health insurance?": "has_health_insurance",
}


class Location(Enum):
    EAST_NY = "CHN Vaccination Site: Church of God (East NY)"
    HARLEM = "CHN Vaccination Site: Convent Baptist (Harlem)"
    WASHINGTON_HEIGHTS = "CHN Vaccination Site: Fort Washington (Washington Heights)"
    SOUTH_JAMAICA = "CHN Vaccination Site: New Jerusalem (South Jamaica)"


class Race(Enum):
    ASIAN = "Asian (including South Asian)"
    BLACK = "Black including African American or Afro-Caribbean"
    NATIVE_AMERICAN = "Native American or Alaska Native"
    WHITE = "White"
    PACIFIC_ISLANDER = "Native Hawaiian or Pacific Islander"
    OTHER = "Other"
    PREFER_NOT_TO_ANSWER = "Prefer not to answer"


class Sex(Enum):
    MALE = "Male"
    FEMALE = "Female"
    NEITHER = "Neither"
    UNKNOWN = "Unknown"


class Ethnicity(Enum):
    LATINX = "Yes"
    NOT_LATINX = "No"
    PERFER_NOT_TO_ANSWER = "Prefer not to answer"


class Config:
    anystr_strip_whitespace = True


@dataclass(config=Config)
class FormEntry:
    id: Optional[int]
    start_time: datetime
    first_name: str
    last_name: str
    phone: Optional[PositiveInt]
    email: EmailStr
    location: Location
    dob: datetime
    street_address: str
    city: str
    state: Literal["NY", "NJ"]
    apt: Optional[str]
    zip_code: PositiveInt
    race: Race
    ethnicity: Ethnicity
    sex: Sex
    has_health_insurance: bool

    @property
    def date_str(self):
        return self.start_time.strftime("%m/%d/%Y")

    @property
    def time_str(self):
        return self.start_time.strftime("%I:%M %p")

    @property
    def dob_str(self):
        return self.dob.strftime("%m/%d/%Y")

    @validator("email")
    def not_empty(cls, v):
        if len(v) == 0:
            raise ValueError("Empty field.")
        return v

    @validator("start_time")
    def strip_tzinfo(cls, dt):
        return dt.replace(tzinfo=None)

    @validator("apt")
    def empty_as_none(cls, v):
        return None if v == "" else v

    @validator("state", pre=True)
    def cast_state(cls, v):
        # No validation for "state" in Acuity.
        # This function safely determines whether
        upper = str(v).strip().upper()
        if upper in {"NJ", "NY"}:
            return upper
        if "YORK" in upper:
            return "NY"
        if "JERSEY" in upper:
            return "NJ"
        return v

    @validator("phone", pre=True)
    def cast_phone(cls, v):
        if isinstance(v, int):
            if len(str(v)) != 10:
                raise ValueError("Phone number is longer than 10 digits")
            return v
        # Since it's not a required field, only provide number if its exactly 10 digits
        v = re.sub(r"[^0-9]+", "", str(v))[-10:]  # remove non-numeric characters
        v = v[-10:]  # and last 10 elements (or less)
        return v if len(v) == 10 else None

    @validator("dob", pre=True)
    def instance_dt(cls, v):
        if isinstance(v, datetime):
            return v
        return datetime.strptime(v.strip(), "%m/%d/%Y")


class AcuityExportReader:
    def __init__(self, path: str):
        wb = load_workbook(path)
        self.name = wb.sheetnames[0]
        self._ws = wb[self.name]

    def __iter__(self):
        data = self._ws.values
        headers = next(data)
        for row in data:
            if all(i is None for i in row):
                # empty row
                continue
            # Rename and pick fields
            yield {COLUMNS[k]: v for k, v in zip(headers, row) if k in COLUMNS}

    def __len__(self):
        data = self._ws.values
        next(data)
        return sum(1 for _ in data)


def group_entries(entries: Iterable[FormEntry]):
    sorted_entries = sorted(entries, key=lambda e: e.location.value)
    return groupby(sorted_entries, key=lambda e: e.location)


DUMMY_DATA = {
    "id": 100000,
    "first_name": "Trevor",
    "last_name": "Manz",
    "start_time": "2021-04-28T21:30",
    "phone": "7158289308",
    "email": "trevmanz94@gmail.com",
    "location": Location.EAST_NY,
    "dob": "07/07/1994",
    "street_address": "500 WN street",
    "city": "New York",
    "state": "NY",
    "apt": None,
    "zip_code": "10001",
    "race": Race.WHITE,
    "ethnicity": Ethnicity.NOT_LATINX,
    "sex": Sex.MALE,
    "has_health_insurance": "yes",
}